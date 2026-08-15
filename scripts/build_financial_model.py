from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

out = Path('/home/ubuntu/duck-hub/docs/modelo_financiero_duck.xlsx')
out.parent.mkdir(parents=True, exist_ok=True)
wb = Workbook()
ws = wb.active
ws.title = 'Supuestos'

headers = ['Variable', 'Conservador', 'Base', 'Expansivo', 'Unidad', 'Origen / nota']
ws.append(headers)
rows = [
    ('Proyectos de producción / mes', 2, 4, 7, 'proyectos', 'Supuesto editable; no es dato histórico'),
    ('Ticket medio de producción', 500, 750, 1000, 'EUR', 'Supuesto editable'),
    ('Beats no exclusivos / mes', 3, 8, 15, 'ventas', 'Supuesto editable'),
    ('Precio medio beat no exclusivo', 60, 75, 90, 'EUR', 'Supuesto editable'),
    ('Beats exclusivos / mes', 0, 1, 2, 'ventas', 'Supuesto editable'),
    ('Precio medio beat exclusivo', 0, 450, 700, 'EUR', 'Supuesto editable'),
    ('Retainers / mes', 0, 1, 3, 'clientes', 'Supuesto editable'),
    ('Precio medio retainer', 0, 600, 900, 'EUR', 'Supuesto editable'),
    ('Coste variable / proyecto', 120, 150, 220, 'EUR', 'Supuesto editable'),
    ('Coste variable / venta de beat', 5, 7, 10, 'EUR', 'Supuesto editable'),
    ('Coste fijo mensual', 250, 450, 850, 'EUR', 'Supuesto editable'),
    ('Marketing y adquisición mensual', 50, 150, 400, 'EUR', 'Supuesto editable'),
    ('Caja inicial', 1500, 3000, 6000, 'EUR', 'Supuesto editable'),
    ('Clientes activos al inicio', 2, 4, 6, 'clientes', 'Supuesto editable'),
    ('Clientes nuevos / mes', 1, 3, 6, 'clientes', 'Supuesto editable'),
    ('Clientes perdidos / mes', 1, 1, 1, 'clientes', 'Supuesto editable'),
]
for r in rows:
    ws.append(r)

for col in range(1, 7):
    ws.cell(1, col).font = Font(bold=True, color='FFFFFF')
    ws.cell(1, col).fill = PatternFill('solid', fgColor='0B6B4F')
    ws.cell(1, col).alignment = Alignment(horizontal='center')
for widths in {'A': 34, 'B': 16, 'C': 16, 'D': 16, 'E': 14, 'F': 42}.items():
    ws.column_dimensions[widths[0]].width = widths[1]

calc = wb.create_sheet('Escenarios')
calc.append(['Métrica', 'Conservador', 'Base', 'Expansivo', 'Interpretación'])
for c in range(1, 6):
    calc.cell(1, c).font = Font(bold=True, color='FFFFFF')
    calc.cell(1, c).fill = PatternFill('solid', fgColor='0B6B4F')
metrics = [
    ('Ingresos producción / mes', '=Supuestos!B2*Supuestos!B3', '=Supuestos!C2*Supuestos!C3', '=Supuestos!D2*Supuestos!D3', 'Volumen por ticket'),
    ('Ingresos beats no exclusivos / mes', '=Supuestos!B4*Supuestos!B5', '=Supuestos!C4*Supuestos!C5', '=Supuestos!D4*Supuestos!D5', 'Catálogo recurrente'),
    ('Ingresos beats exclusivos / mes', '=Supuestos!B6*Supuestos!B7', '=Supuestos!C6*Supuestos!C7', '=Supuestos!D6*Supuestos!D7', 'Escasez y licencia'),
    ('Ingresos retainers / mes', '=Supuestos!B8*Supuestos!B9', '=Supuestos!C8*Supuestos!C9', '=Supuestos!D8*Supuestos!D9', 'Previsibilidad'),
    ('Ingresos totales / mes', '=SUM(B2:B5)', '=SUM(C2:C5)', '=SUM(D2:D5)', 'No equivale a caja cobrada'),
    ('Costes variables / mes', '=Supuestos!B2*Supuestos!B10+Supuestos!B4*Supuestos!B11+Supuestos!B6*Supuestos!B11', '=Supuestos!C2*Supuestos!C10+Supuestos!C4*Supuestos!C11+Supuestos!C6*Supuestos!C11', '=Supuestos!D2*Supuestos!D10+Supuestos!D4*Supuestos!D11+Supuestos!D6*Supuestos!D11', 'Antes de costes fijos'),
    ('Margen de contribución / mes', '=B6-B7', '=C6-C7', '=D6-D7', 'Ingresos menos variables'),
    ('Costes fijos + marketing / mes', '=Supuestos!B12+Supuestos!B13', '=Supuestos!C12+Supuestos!C13', '=Supuestos!D12+Supuestos!D13', 'Gasto operativo'),
    ('Resultado operativo aproximado / mes', '=B8-B9', '=C8-C9', '=D8-D9', 'No es beneficio fiscal'),
    ('Punto de equilibrio de ingresos', '=B9/(1-B7/B6)', '=C9/(1-C7/C6)', '=D9/(1-D7/D6)', 'Ingresos aproximados'),
    ('Runway en meses si resultado negativo', '=IF(B10<0,Supuestos!B14/MAX(1,-B10),NA())', '=IF(C10<0,Supuestos!C14/MAX(1,-C10),NA())', '=IF(D10<0,Supuestos!D14/MAX(1,-D10),NA())', 'Caja / pérdida mensual'),
]
for r in metrics:
    calc.append(r)
for col in range(1, 6):
    calc.column_dimensions[get_column_letter(col)].width = [38, 18, 18, 18, 38][col-1]

monthly = wb.create_sheet('Caja 12 meses')
monthly.append(['Mes', 'Caja conservadora', 'Caja base', 'Caja expansiva', 'Clientes activos base'])
for c in range(1, 6):
    monthly.cell(1, c).font = Font(bold=True, color='FFFFFF')
    monthly.cell(1, c).fill = PatternFill('solid', fgColor='0B6B4F')
for m in range(1, 13):
    row = m + 1
    monthly.cell(row, 1, m)
    if m == 1:
        monthly.cell(row, 2, '=Supuestos!B14+Escenarios!B10')
        monthly.cell(row, 3, '=Supuestos!C14+Escenarios!C10')
        monthly.cell(row, 4, '=Supuestos!D14+Escenarios!D10')
        monthly.cell(row, 5, '=MAX(0,Supuestos!C15+Supuestos!C16-Supuestos!C17)')
    else:
        monthly.cell(row, 2, f'=B{row-1}+Escenarios!B10')
        monthly.cell(row, 3, f'=C{row-1}+Escenarios!C10')
        monthly.cell(row, 4, f'=D{row-1}+Escenarios!D10')
        monthly.cell(row, 5, f'=MAX(0,E{row-1}+Supuestos!C16-Supuestos!C17)')
for col, width in {'A':12, 'B':20, 'C':20, 'D':20, 'E':22}.items():
    monthly.column_dimensions[col].width = width

risk = wb.create_sheet('Riesgos y métricas')
risk.append(['Riesgo / métrica', 'Umbral de alerta', 'Acción'])
for c in range(1, 4):
    risk.cell(1, c).font = Font(bold=True, color='FFFFFF')
    risk.cell(1, c).fill = PatternFill('solid', fgColor='0B6B4F')
for r in [
    ('Concentración del mayor cliente', '> 35% de ingresos', 'Diversificar antes de aumentar costes fijos'),
    ('Caja proyectada', '< 3 meses de costes fijos', 'Reducir gasto y priorizar cobros'),
    ('Margen de contribución', '< 50%', 'Revisar ticket, alcance y costes'),
    ('Revisiones extra', '> 20% de proyectos', 'Reforzar contrato y límite server-side'),
    ('Pagos atrasados', '> 15% de facturas', 'Cobro anticipado o hitos'),
    ('Dependencia de un canal', '> 60% de leads', 'Abrir canal directo y catálogo propio'),
]:
    risk.append(r)
for col, width in {'A':32, 'B':26, 'C':56}.items():
    risk.column_dimensions[col].width = width

for sheet in wb.worksheets:
    sheet.freeze_panes = 'A2'
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

wb.save(out)
print(out)
